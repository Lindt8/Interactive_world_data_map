'''
Created on Mar 10, 2021
'''

import codecs
import country_converter as coco    # pip install country_converter --upgrade
import pygal                        # pip install pygal
from pygal.maps.world import World  # pip install pygal_maps_world
from pygal.style import Style
import openpyxl
import datetime
import numpy as np
import os
import sys 
import PySimpleGUI as sg   # pip install pysimplegui
import tkinter as tk
import pickle
import csv
import string

def fix_map_country_names(map_filepath):
    f = open(map_filepath, mode='r', encoding='utf_8')
    new_text = ''
    for line in f:
        text = line
        if '<desc class="value">' in text:
            part1,part2 = text.split(':',1)
            country_name_old = part1.replace('<desc class="value">','').strip()
            country_name_new = coco.convert(country_name_old, to='short_name')
            text = '    <desc class="value">'+country_name_new+':'+part2
            #for ni in range(len(pygal_country_full_names)):
            #    if pygal_country_full_names[ni] in line: 
            #        #print(pygal_country_full_names[ni])
            #        text = line.replace(pygal_country_full_names[ni],coco.convert(pygal_country_full_names[ni], to='short_name'))
            #        break
            new_text += text
        else:
            new_text += text
    f.close()
    # Fix strange "ns1:" additions on Japanese systems
    new_text = new_text.replace("ns1:","") 
    f2 = open(map_filepath,"w+", encoding='utf_8')
    f2.write(new_text)
    f2.close() 


def add_count_bins_to_map(worldmap_chart,count_bins=[1,10,25,50,100]):
    global include_overflow_bin, force_integer_bin_edges
    nbins=len(count_bins)
    if not include_overflow_bin:
        nbins = nbins - 1
    for i in range(nbins):
        if i==nbins-1 and include_overflow_bin: # overflow bin
            if force_integer_bin_edges:
                worldmap_chart.add('{:g}+'.format(count_bins[i]+1), [])
            else:
                worldmap_chart.add('{:g}+'.format(count_bins[i]), [])
        else:
            if i==0:
                left_val = count_bins[i]
            else:
                if force_integer_bin_edges:
                    left_val = count_bins[i]+1
                else:
                    left_val = count_bins[i]
            if force_integer_bin_edges and count_bins[i+1]==left_val:
                worldmap_chart.add('{:g}'.format(left_val), [])
            elif force_integer_bin_edges and left_val==0 and count_bins[i+1]==1:
                worldmap_chart.add('{:g}'.format(count_bins[i+1]), [])
            else:
                worldmap_chart.add('{:g}-{:g}'.format(left_val,count_bins[i+1]), [])
    return worldmap_chart

def generate_count_bins(data,nbins=4,spacing_style='lin',force_integer_bin_edges=True):
    global include_overflow_bin
    count_bins = []
    n_edges = nbins + 1
    if include_overflow_bin: 
        n_edges += 1
    min_value = min([i for i in data.values()])
    max_value = max([i for i in data.values()])
    if spacing_style=='log':
        min_value = np.log10(min_value) 
        max_value = np.log10(max_value) 
    if include_overflow_bin:
        max_value = max_value*(n_edges-1)/(n_edges)
    if force_integer_bin_edges:
        bin_dtype = np.int
    else:
        bin_dtype = None
    if spacing_style=='lin':
        count_bins = np.linspace(min_value,max_value,num=nbins+1,dtype=bin_dtype)
    else: # 'log'
        count_bins = np.logspace(min_value,max_value,num=nbins+1,dtype=bin_dtype)
    if force_integer_bin_edges: 
        count_bins[-1] += 1
        
    return count_bins
    
def bin_map_colors(map_filepath,data,count_bins=[1,10,25,50,100],min_opacity=0.15,max_opacity=0.95):
    global legend_label, show_legend, map_hover_text, include_overflow_bin, force_integer_bin_edges
    f = open(map_filepath, mode='r', encoding='utf_8')
    new_text = ''
    nbins = len(count_bins) # including overflow bin
    if not include_overflow_bin: nbins = nbins-1
    opacity_bins = np.linspace(min_opacity,max_opacity,num=nbins).tolist()
    #opacity_bins += [1.0]
    ln = 0
    for line in f:
        ln += 1
        text = line
        if ln == 4 and show_legend:  # Fix label shown in hover box
            pre_legend, post_legend = text.split('</script>')
            pre_legend = pre_legend.split('"legends"')[0]
            new_line = ''
            new_line += pre_legend + '"legends": ['
            for i in range(nbins+1):
                new_line += '"{}"'.format(map_hover_text)
                if i!=nbins:
                    new_line += ', '
            new_line += ']}</script>' + post_legend
            new_text += new_line
        elif 'country map-element color-0' in text: 
            for key in data:
                if key in line.split('"')[1][:2]: 
                    if force_integer_bin_edges:
                        if data[key]<=count_bins[0]:
                            oi = 0
                            opacity = opacity_bins[0]
                        elif include_overflow_bin and data[key]>count_bins[-1]:
                            oi = len(opacity_bins)-1
                            opacity = opacity_bins[-1]
                        elif not include_overflow_bin and data[key]>count_bins[-1]:
                            opacity = 0.0
                        else:
                            #oi=next(i for i,v in enumerate(count_bins) if v >= data[key])-1
                            oi = 0
                            for oiii in range(len(count_bins)):
                                if data[key] <= count_bins[oiii]: break
                                oi = oiii
                            opacity=opacity_bins[oi]
                    else:
                        if data[key]<=count_bins[0]:
                            oi = 0
                            opacity = opacity_bins[0]
                        elif include_overflow_bin and data[key]>=count_bins[-1]:
                            oi = len(opacity_bins)-1
                            opacity = opacity_bins[-1]
                        elif not include_overflow_bin and data[key]>count_bins[-1]:
                            opacity = 0.0
                        else:
                            #oi=next(i for i,v in enumerate(count_bins) if v >= data[key])-1
                            oi = 0
                            for oiii in range(len(count_bins)):
                                if include_overflow_bin:
                                    if data[key] < count_bins[oiii]: break
                                else:
                                    if data[key] <= count_bins[oiii]: break
                                oi = oiii
                            opacity=opacity_bins[oi]
                    text =     '<g class="{} country map-element color-0 serie-{} series" style="fill-opacity: {:.6f}">'.format(key,oi+1,opacity)
                    break
            new_text += text
        else:
            new_text += text
    f.close()
    
    if show_legend:
        # Fix legend colors
        body, last_line = new_text.rsplit('\n',1)
        for i in range(nbins):
            last_line = last_line.replace('class="color-{} reactive"'.format(str(i+1)),'class="color-0 reactive" style="fill-opacity: {:.6f}"'.format(opacity_bins[i]))
        
        # Label for vertical legend
        # last_line = last_line.replace(tallied_quantity,legend_label)  
        # Label for horizontal legend
        segs = last_line.split('</text>')
        last_line=''
        for i in range(len(segs)):
            if i==len(segs)-1: 
                last_line += segs[i]
            elif i==1:
                main_seg, problem_part = segs[i].rsplit('>',1)
                last_line += main_seg + '>' + legend_label + '</text>'
            else:
                last_line += segs[i] + '</text>'
        new_text = body + last_line
    
    f2 = open(map_filepath,"w+", encoding='utf_8')
    f2.write(new_text)
    f2.close() 

def write_country_tables(data,count_title,left_col_title='Country'):
    import random, string
    x = ''.join(random.choices(string.ascii_letters + string.digits, k=16))
    tab_text = ''
    tab_text += '<style scoped="scoped">' + '\n'
    tab_text += '                #table-{}'.format(x)+' {' + '\n'
    tab_text += '                    border-collapse: collapse;' + '\n'
    tab_text += '                    border-spacing: 0;' + '\n'
    tab_text += '                    empty-cells: show;' + '\n'
    tab_text += '                    border: 1px solid #cbcbcb;' + '\n'
    tab_text += '                    text-align: left;' + '\n'
    tab_text += '                    margin-left:auto;' + '\n'
    tab_text += '                    margin-right:auto;' + '\n'
    tab_text += '                }' + '\n'
    tab_text += '                #table-{}'.format(x)+' td, #table-{}'.format(x)+' th {' + '\n'
    tab_text += '                    border-left: 1px solid #cbcbcb;' + '\n'
    tab_text += '                    border-width: 0 0 0 1px;' + '\n'
    tab_text += '                    margin: 0;' + '\n'
    tab_text += '                    padding: 0.5em 1em;' + '\n'
    tab_text += '                }' + '\n'
    tab_text += '                #table-{}'.format(x)+' td:first-child, #table-{}'.format(x)+' th:first-child {' + '\n'
    tab_text += '                    border-left-width: 0;' + '\n'
    tab_text += '                }' + '\n'
    tab_text += '                #table-{}'.format(x)+' thead, #table-{}'.format(x)+' tfoot {' + '\n'
    tab_text += '                    color: #000;' + '\n'
    tab_text += '                    text-align: left;' + '\n'
    tab_text += '                    vertical-align: bottom;' + '\n'
    tab_text += '                }' + '\n'
    tab_text += '                #table-{}'.format(x)+' thead {' + '\n'
    tab_text += '                    background: #e0e0e0;' + '\n' + '\n'
    tab_text += '                }' + '\n'
    tab_text += '                #table-{}'.format(x)+' tfoot {' + '\n'
    tab_text += '                    background: #ededed;' + '\n'
    tab_text += '                }' + '\n'
    tab_text += '                #table-{}'.format(x)+' tr:nth-child(2n-1) td {' + '\n'
    tab_text += '                    background-color: #f2f2f2;' + '\n'
    tab_text += '                }' + '\n'
    tab_text += '   </style><table id="table-{}"><thead><tr><th>{}</th><th>{}</th></tr></thead><tbody>'.format(x,left_col_title,count_title)
    running_sum = 0
    for key in data:
        tab_text += '<tr><td>{}</td><td>{:g}</td></tr>'.format(coco.convert(key, to='short_name'),data[key])
        running_sum += data[key]
    tab_text += '<tr><td><b>Total</b></td><td><b>{:g}</b></td></tr>'.format(running_sum)
    return tab_text






# If True, below user values are used as defaults in the GUI.
# If False, the GUI is not used at all, and the below values are used
USE_GUI = True


# ========================================================================
#                START OF USER CONTROLLED PARAMETERS
# ========================================================================

# Search for default folder and filepath
# Get default path to file
try:
    f=open('default_path_to_data_file.txt')
    lines=f.readlines()
    f.close()
    data_filepath = lines[0].strip()
except:
    data_filepath = os.getcwd() + '/' +  'test_data/olympics_statistics.xlsx'
if '.xls' in data_filepath:
    data_is_in_spreadsheet = True
else:
    data_is_in_spreadsheet = False
    if os.path.splitext(data_filepath)[1] not in ['.csv','.tsv','.txt']:
        print('Selected data file format unknown, will assume it is a .csv file.')
        filetype = 'csv'
    elif os.path.splitext(data_filepath)[1] == '.tsv':
        filetype = 'tsv'
    else:
        filetype = 'csv'
data_folder, data_filename = os.path.split(data_filepath) #+ '/'
output_folder = data_folder + '/' 

# Data file can be CSV, TSV, or Excel format
i_sheet = 0
wbname = '' # if specified, this overrides i_sheet
n_skip_rows = 1
i_country_col = '1'
is_data_already_in_tallied_form = True
i_data_col = '2' # if is_data_already_in_tallied_form==True, this tells what column is associated with each country's data value


save_data_to_pickle = True # universal toggle for whether data is saved to pickle
read_data_from_pickle = True # toggle whether code will attempt to read pickle; will be set to False if file not found
pickle_filepath = os.path.splitext(data_filepath)[0] + '.pickle'

alternate_tables_and_maps = True # If false, all maps printed first followed by tables
html_include_map_style_options = ['embed', 'object', 'img', 'iframe']
html_include_map_style = html_include_map_style_options[0] # select 'img', 'embed', 'object', 'iframe'

# specify color bins to use
# For example, [1,10,25,50,100] results in separate colors (bins) for values of 1-10, 11-25, 26-50, 51-100, and 101+
manually_enter_bins = True
automatically_generate_bins_lin = False
automatically_generate_bins_log = False
num_auto_bins = 4
tally_bin_edges =     [1,10,25,50,100,200,500] 
include_overflow_bin = True
force_integer_bin_edges = True

show_legend = True 
legend_label = 'Count:'

custom_style_color = '#0000FF'


user_tallied_quantity = 'Tallied quantity'          # this will appear in the table and, when the legend is disabled, the hover box in the SVG when mousing over countries
user_table_coulmn_header = user_tallied_quantity    # this is the header in the second column of the output table
user_table_country_coulmn_header = 'Country'        # this is the header in the first column of the output table
user_map_hover_text = user_tallied_quantity + ' in' # this will appear in the hover box in the SVG when mousing over countries when the legend is enabled
user_map_title = 'Number of X by country'
user_map_output_filename = 'Value_worldmap'
user_html_output_filename = user_map_output_filename


# ========================================================================
#                 END OF USER CONTROLLED PARAMETERS
# ========================================================================




# ========================================================================
#                START OF THE ACTUAL PROGRAM
# ========================================================================

def generate_maps_and_tables():
    global legend_label, show_legend, map_hover_text, read_data_from_pickle, wbname, i_sheet
    
    custom_style = Style(colors=(custom_style_color, custom_style_color))
    
    # ISO2 names used by Pygal (only includes countries drawn to maps)
    ISO2_country_names = ["af","al","dz","ad","ao","aq","ar","am","au","at","az","bh","bd","by","be","bz","bj","bt","bo","ba","bw","br","bn","bg","bf","bi","kh","cm","ca","cv","cf","td","cl","cn","co","cg","cd","cr","ci","hr","cu","cy","cz","dk","dj","do","ec","eg","sv","gq","er","ee","et","fi","fr","gf","ga","gm","ge","de","gh","gr","gl","gu","gt","gn","gw","gy","ht","va","hn","hk","hu","is","in","id","ir","iq","ie","il","it","jm","jp","jo","kz","ke","kp","kr","kw","kg","la","lv","lb","ls","lr","ly","li","lt","lu","mo","mk","mg","mw","my","mv","ml","mt","mr","mu","yt","mx","md","mc","mn","me","ma","mz","mm","na","np","nl","nz","ni","ne","ng","no","om","pk","ps","pa","pg","py","pe","ph","pl","pt","pr","re","ro","ru","rw","sh","sm","st","sa","sn","rs","sc","sl","sg","sk","si","so","za","es","lk","sd","sr","sz","se","ch","sy","tw","tj","tz","th","tl","tg","tn","tr","tm","ug","ua","ae","gb","us","uy","uz","ve","vn","eh","ye","zm","zw"]
    cnames_checked = [] # this will be used later for debugging
    
    # Pygal country names; these will be replaced by short names.
    pygal_country_full_names = ["Andorra","United Arab Emirates","Afghanistan","Albania","Armenia","Angola","Antarctica","Argentina","Austria","Australia","Azerbaijan","Bosnia and Herzegovina","Bangladesh","Belgium","Burkina Faso","Bulgaria","Bahrain","Burundi","Benin","Brunei Darussalam","Bolivia, Plurinational State of","Brazil","Bhutan","Botswana","Belarus","Belize","Canada","Congo, the Democratic Republic of the","Central African Republic","Congo","Switzerland","Cote d'Ivoire","Chile","Cameroon","China","Colombia","Costa Rica","Cuba","Cape Verde","Cyprus","Czech Republic","Germany","Djibouti","Denmark","Dominican Republic","Algeria","Ecuador","Estonia","Egypt","Western Sahara","Eritrea","Spain","Ethiopia","Finland","France","Gabon","United Kingdom","Georgia","French Guiana","Ghana","Greenland","Gambia","Guinea","Equatorial Guinea","Greece","Guatemala","Guam","Guinea-Bissau","Guyana","Hong Kong","Honduras","Croatia","Haiti","Hungary","Indonesia","Ireland","Israel","India","Iraq","Iran, Islamic Republic of","Iceland","Italy","Jamaica","Jordan","Japan","Kenya","Kyrgyzstan","Cambodia","Korea, Democratic People's Republic of","Korea, Republic of","Kuwait","Kazakhstan","Lao People's Democratic Republic","Lebanon","Liechtenstein","Sri Lanka","Liberia","Lesotho","Lithuania","Luxembourg","Latvia","Libyan Arab Jamahiriya","Morocco","Monaco","Moldova, Republic of","Montenegro","Madagascar","Macedonia, the former Yugoslav Republic of","Mali","Myanmar","Mongolia","Macao","Mauritania","Malta","Mauritius","Maldives","Malawi","Mexico","Malaysia","Mozambique","Namibia","Niger","Nigeria","Nicaragua","Netherlands","Norway","Nepal","New Zealand","Oman","Panama","Peru","Papua New Guinea","Philippines","Pakistan","Poland","Puerto Rico","Palestine, State of","Portugal","Paraguay","Reunion","Romania","Serbia","Russian Federation","Rwanda","Saudi Arabia","Seychelles","Sudan","Sweden","Singapore","Saint Helena, Ascension and Tristan da Cunha","Slovenia","Slovakia","Sierra Leone","San Marino","Senegal","Somalia","Suriname","Sao Tome and Principe","El Salvador","Syrian Arab Republic","Swaziland","Chad","Togo","Thailand","Tajikistan","Timor-Leste","Turkmenistan","Tunisia","Turkey","Taiwan (Republic of China)","Taiwan, Province of China","Tanzania, United Republic of","Ukraine","Uganda","United States","Uruguay","Uzbekistan","Holy See (Vatican City State)","Venezuela, Bolivarian Republic of","Viet Nam","Yemen","Mayotte","South Africa","Zambia","Zimbabwe"]
    
    '''
    # Temporary code used to determine countries not in Pygal but in coco
    not_in_pygal_list = []
    coco_list = ["Afghanistan","Aland Islands","Albania","Algeria","American Samoa","Andorra","Angola","Anguilla","Antarctica","Antigua and Barbuda","Argentina","Armenia","Aruba","Australia","Austria","Azerbaijan","Bahamas","Bahrain","Bangladesh","Barbados","Belarus","Belgium","Belize","Benin","Bermuda","Bhutan","Bolivia","Bonaire, Saint Eustatius and Saba","Bosnia and Herzegovina","Botswana","Bouvet Island","Brazil","British Antarctic Territories","British Indian Ocean Territory","British Virgin Islands","Brunei Darussalam","Bulgaria","Burkina Faso","Burundi","Cabo Verde","Cambodia","Cameroon","Canada","Cayman Islands","Central African Republic","Chad","Channel Islands","Chile","China","Christmas Island","Cocos (Keeling) Islands","Colombia","Comoros","Congo Republic","Cook Islands","Costa Rica","Cote d'Ivoire","Croatia","Cuba","Curacao","Cyprus","Czech Republic","Denmark","Djibouti","Dominica","Dominican Republic","DR Congo","Ecuador","Egypt","El Salvador","Equatorial Guinea","Eritrea","Estonia","Eswatini","Ethiopia","Faeroe Islands","Falkland Islands","Fiji","Finland","France","French Guiana","French Polynesia","French Southern Territories","Gabon","Gambia","Georgia","Germany","Ghana","Gibraltar","Greece","Greenland","Grenada","Guadeloupe","Guam","Guatemala","Guernsey","Guinea","Guinea-Bissau","Guyana","Haiti","Heard and McDonald Islands","Honduras","Hong Kong","Hungary","Iceland","India","Indonesia","Iran","Iraq","Ireland","Isle of Man","Israel","Italy","Jamaica","Japan","Jersey","Jordan","Kazakhstan","Kenya","Kiribati","Kosovo","Kuwait","Kyrgyz Republic","Laos","Latvia","Lebanon","Lesotho","Liberia","Libya","Liechtenstein","Lithuania","Luxembourg","Macau","Macedonia","Madagascar","Malawi","Malaysia","Maldives","Mali","Malta","Marshall Islands","Martinique","Mauritania","Mauritius","Mayotte","Mexico","Micronesia, Fed. Sts.","Moldova","Monaco","Mongolia","Montenegro","Montserrat","Morocco","Mozambique","Myanmar","Namibia","Nauru","Nepal","Netherlands","Netherlands Antilles","New Caledonia","New Zealand","Nicaragua","Niger","Nigeria","Niue","Norfolk Island","North Korea","Northern Mariana Islands","Norway","Oman","Pakistan","Palau","Palestine","Panama","Papua New Guinea","Paraguay","Peru","Philippines","Pitcairn","Poland","Portugal","Puerto Rico","Qatar","Reunion","Romania","Russia","Rwanda","Saint-Martin","Samoa","San Marino","Sao Tome and Principe","Saudi Arabia","Senegal","Serbia","Seychelles","Sierra Leone","Singapore","Sint Maarten","Slovakia","Slovenia","Solomon Islands","Somalia","South Africa","South Georgia and South Sandwich Is.","South Korea","South Sudan","Soviet Union (former)","Spain","Sri Lanka","St. Barths","St. Helena","St. Kitts and Nevis","St. Lucia","St. Pierre and Miquelon","St. Vincent and the Grenadines","Sudan","Suriname","Svalbard and Jan Mayen Islands","Sweden","Switzerland","Syria","Taiwan","Tajikistan","Tanganjika","Tanzania","Thailand","Timor-Leste","Togo","Tokelau","Tonga","Trinidad and Tobago","Tunisia","Turkey","Turkmenistan","Turks and Caicos Islands","Tuvalu","Uganda","Ukraine","United Arab Emirates","United Kingdom","United States","United States Minor Outlying Islands","United States Virgin Islands","Uruguay","Uzbekistan","Vanuatu","Vatican","Venezuela","Vietnam","Wallis and Futuna Islands","Western Sahara","Yemen","Zambia","Zanzibar","Zimbabwe"]
    for i in coco_list:
        x = coco.convert(i, to='ISO2')
        if x.lower() not in ISO2_country_names and x!='not found':
            not_in_pygal_list.append(x)
    print(not_in_pygal_list)
    sys.exit()
    '''
    
    # Exceptional country names to auto-correct when read in
    pre_correction_read_country_names  = ["UK"            , "West Germany", "East Germany", "Soviet Union"]
    post_correction_read_country_names = ["United Kingdom", "Germany"     , "Germany"     , "Russia"      ]
    
    # ISO2 codes of real countries not available in Pygal (mostly due to being too small to appear on map)
    pygal_unsupported_countries = ['AX', 'AS', 'AI', 'AG', 'AW', 'BS', 'BB', 'BM', 'BQ', 'BV', 'IO', 'VG', 'KY', 'CX', 'CC', 'KM', 'CK', 'CW', 'DM', 'FO', 'FK', 'FJ', 'PF', 'TF', 'GI', 'GD', 'GP', 'GG', 'HM', 'IM', 'JE', 'KI', 'XK', 'MH', 'MQ', 'FM', 'MS', 'NR', 'NC', 'NU', 'NF', 'MP', 'PW', 'PN', 'QA', 'MF', 'WS', 'SX', 'SB', 'GS', 'SS', 'BL', 'KN', 'LC', 'PM', 'VC', 'SJ', 'TK', 'TO', 'TT', 'TC', 'TV', 'UM', 'VI', 'VU', 'WF']
    
    html_item_spacer = '\n\n<br>\n\n'
    
    #'img', 'embed', 'object', 'iframe'
    if html_include_map_style=='object':
        svg_html_code = '<div style="text-align: center;"><object data="{}" type="image/svg+xml" style="width:50%;"></object></div>'
    elif html_include_map_style=='img':
        svg_html_code = '<div style="text-align: center;"><img src="{}" type="image/svg+xml" style="width:50%;"></object></div>'
    elif html_include_map_style=='embed':
        svg_html_code = '<div style="text-align: center;"><embed src="{}" type="image/svg+xml" style="width:50%;"></object></div>'
    elif html_include_map_style=='iframe':
        svg_html_code = '<div style="text-align: center;"><iframe src="{}" type="image/svg+xml" frameborder="0" scrolling="no" height=700px style="width:50%;"></iframe></object></div>'
    
    start_time = datetime.datetime.now()
    
    user_data_pickle = {}
    if os.path.isfile(pickle_filepath):
        if read_data_from_pickle:
            with open(pickle_filepath, 'rb') as handle:
                user_data_pickle = pickle.load(handle)
    else:
        read_data_from_pickle = False
    
    
    the_spreadsheet_was_parsed = False
    file = data_filepath
    datetime_last_modified = datetime.datetime.fromtimestamp(os.path.getmtime(file))
    if data_is_in_spreadsheet:
        wb = openpyxl.load_workbook(filename = file, read_only=True,data_only=True)  # read only is important for making it load quickly
    #else:
    #    f = open(file)
    #    file_text = f.readlines()
    #    f.close()
    print('loading file: '+repr(file)+'\n\t{} elapsed'.format(datetime.datetime.now() - start_time))
    
    html_output_text = ''
    html_maps_output_text = ''
    html_tables_output_text = ''

    tallied_quantity = user_tallied_quantity  # this will appear in the table and, when the legend is disabled, the hover box in the SVG when mousing over countries
    table_coulmn_header = user_table_coulmn_header # this is the header in the second column of the output table
    table_country_coulmn_header = user_table_country_coulmn_header # header of the table's first column
    map_hover_text = user_map_hover_text # this will appear in the hover box in the SVG when mousing over countries when the legend is enabled
    map_title = user_map_title
    map_output_filename = user_map_output_filename
    html_output_filename = user_html_output_filename
    
    if data_is_in_spreadsheet:
        if USE_GUI: #i_sheet = wb.sheetnames.index(wbname)
            map_category = wbname
        else:
            map_category = wb.sheetnames[i_sheet]
    else:
        map_category = ''

    
    map_output_filepath = output_folder + map_output_filename + '.svg'
    html_output_filepath = output_folder + html_output_filename + '.html'
    
    
    
    # Determine if data from Excel is already stored in pickle file
    main_data_keyname = map_category + '_data_all'
    read_THIS_data_from_pickle = False
    if read_data_from_pickle:
        if main_data_keyname in user_data_pickle:
            if user_data_pickle[main_data_keyname]['datetime_last_modified'] == datetime_last_modified:
                read_THIS_data_from_pickle = True
    
    
    if read_THIS_data_from_pickle:
        print('found up-to-date {} data in pickle file; using it instead of data file (substantially faster)...\n\t{} elapsed'.format(map_category,datetime.datetime.now() - start_time))
        main_tally = user_data_pickle[main_data_keyname]['data']
    
    
    else: # have to actually parse spreadsheet for data
        the_spreadsheet_was_parsed = True
        # Initialize dictionaries to tally users
        main_tally = {}
        for ci in range(len(ISO2_country_names)):
            main_tally.update({ISO2_country_names[ci] : 0})
        
        if data_is_in_spreadsheet:
            if wbname=='':
                user_sheet = wb.worksheets[i_sheet]
                wbname = wb.sheetnames[i_sheet]
            else:
                i_sheet = wb.sheetnames.index(wbname)
                user_sheet = wb.worksheets[i_sheet]
            
            
            # Extract user data
            if cnames_checked==[]: print('Input country name, discovered ISO2 country code, and corresponding output country names are printed the first time they are encountered.  Please check if any countries are improperly identified.')
            print('extracting statistics from the "{}" sheet...\n\t{} elapsed'.format(wbname,datetime.datetime.now() - start_time))
                
            last_row_num = user_sheet.max_row
            print('\tfound {} rows'.format(last_row_num))
            rn = 0
            
            col_input = i_country_col
            try:
                col_letter_country = ''
                col_num = int(col_input)
                if col_num==0: col_num=1
                if col_num>26:
                    col_letter_country += chr(ord('@')+col_num//26)
                    col_num = col_num%26
                col_letter_country += chr(ord('@')+col_num)
            except:
                col_letter_country = col_input.upper()
            
            if is_data_already_in_tallied_form:
                col_input = i_data_col
                try:
                    col_letter_data = ''
                    col_num = int(col_input)
                    if col_num==0: col_num=1
                    if col_num>26:
                        col_letter_data += chr(ord('@')+col_num//26)
                        col_num = col_num%26
                    col_letter_data += chr(ord('@')+col_num)
                except:
                    col_letter_data = col_input.upper()
            
            
            while rn < last_row_num:
                rn += 1
                if rn<=n_skip_rows: continue # skip header rows
                if rn%100==0: print('\t\treached row {}'.format(str(rn)))
                #print(rn)
                
                country_name = user_sheet[col_letter_country+str(rn)].value
                country_name_unmodified = country_name
                if not country_name: continue # skip lines missing countries
                # correct names which normally can't be identified correctly
                for dni, difficult_name in enumerate(pre_correction_read_country_names):
                    if country_name == difficult_name: country_name = post_correction_read_country_names[dni]
                iso2cc = coco.convert(country_name, to='ISO2')
                iso2ccl = iso2cc.lower()
                if country_name not in cnames_checked:
                    print('\t{} --> \t{} --> \t{}'.format(country_name_unmodified,iso2cc,coco.convert(iso2cc, to='short_name')))
                    cnames_checked.append(country_name)
                if iso2cc == 'not found':
                    print('\t\tWarning: skipping since "{}" in row {} could not be properly identified'.format(country_name_unmodified,rn))
                    continue
                if iso2cc in pygal_unsupported_countries:
                    print('\t\tWarning: skipping since {} is not supported by Pygal\n\t\t\t\t(land area too small for map)'.format(coco.convert(iso2cc, to='short_name')))
                    continue
                if is_data_already_in_tallied_form:
                    main_tally[iso2ccl] += user_sheet[col_letter_data+str(rn)].value
                else:
                    main_tally[iso2ccl] += 1
        
        else: # data is not spreadsheet
            print('extracting statistics from the file...\n\t{} elapsed'.format(datetime.datetime.now() - start_time))
            if filetype == 'csv':
                delim = ','
            else: # tsv
                delim = '\t'
             
            # get column index with countries    
            try:
                ci_country = int(i_country_col)-1
            except: # convert from letter
                ci_country = 0
                for li, letter in enumerate(i_country_col[::-1]): # go in reverse order
                    ci_country += (string.ascii_lowercase.index(letter.lower()))*(26**li)
            
            # get column index with data values
            if is_data_already_in_tallied_form:
                try:
                    ci_data = int(i_data_col)-1
                except: # convert from letter
                    ci_data = 0
                    for li, letter in enumerate(i_data_col[::-1]): # go in reverse order
                        ci_data += (string.ascii_lowercase.index(letter.lower()))*(26**li)
            
            with open(data_filepath, mode='r',encoding='unicode_escape') as csv_file:
                spamreader = csv.reader(csv_file,delimiter=delim,quotechar='"')
                rn = 0
                for row in spamreader:
                    rn += 1
                    if rn<=n_skip_rows: continue # skip header rows
                    if rn%100==0: print('\t\treached row {}'.format(str(rn)))
                    
                    country_name = row[ci_country]
                    country_name_unmodified = country_name
                    if not country_name: continue # skip lines missing countries
                    # correct names which normally can't be identified correctly
                    for dni, difficult_name in enumerate(pre_correction_read_country_names):
                        if country_name == difficult_name: country_name = post_correction_read_country_names[dni]
                    iso2cc = coco.convert(country_name, to='ISO2')
                    iso2ccl = iso2cc.lower()
                    if country_name not in cnames_checked:
                        print('\t{} --> \t{} --> \t{}'.format(country_name_unmodified,iso2cc,coco.convert(iso2cc, to='short_name')))
                        cnames_checked.append(country_name)
                    if iso2cc == 'not found':
                        print('\t\tWarning: skipping since "{}" in row {} could not be properly identified'.format(country_name_unmodified,rn))
                        continue
                    if iso2cc in pygal_unsupported_countries:
                        print('\t\tWarning: skipping since {} is not supported by Pygal\n\t\t\t\t(land area too small for map)'.format(coco.convert(iso2cc, to='short_name')))
                        continue
                    if is_data_already_in_tallied_form:
                        main_tally[iso2ccl] += np.float(row[ci_data].replace(',',''))
                    else:
                        main_tally[iso2ccl] += 1
        
        # purge countries with no users
        for cc in ISO2_country_names:
            if main_tally[cc]==0: del main_tally[cc]
        
        # update data in pickle dictionary
        if save_data_to_pickle:
            # Determine if data from Excel is already stored in pickle dictionary file
            if main_data_keyname in user_data_pickle: # overwrite old data
                user_data_pickle[main_data_keyname] = {'datetime_last_modified':datetime_last_modified,'data':main_tally}
            else: # new entry
                user_data_pickle.update({main_data_keyname:{'datetime_last_modified':datetime_last_modified,'data':main_tally}})
    
    print('saving {} map and table... '.format(map_category)+'\n\t{} elapsed'.format(datetime.datetime.now() - start_time))
    
    # Write map SVG file and HTML files with table
    if manually_enter_bins:
        count_bin_edges = tally_bin_edges
    else:
        if automatically_generate_bins_lin:
            spacing_style = 'lin'
        else:
            spacing_style = 'log'
        count_bin_edges = generate_count_bins(main_tally,nbins=num_auto_bins, spacing_style=spacing_style,force_integer_bin_edges=force_integer_bin_edges)
    worldmap_chart = World(show_legend=show_legend,style=custom_style,legend_at_bottom=True, legend_at_bottom_columns=len(count_bin_edges)+1)
    worldmap_chart.title = map_title
    worldmap_chart.add(tallied_quantity, main_tally)
    worldmap_chart = add_count_bins_to_map(worldmap_chart,count_bins=count_bin_edges)
    worldmap_chart.render_to_file(map_output_filepath)
    fix_map_country_names(map_output_filepath)
    bin_map_colors(map_output_filepath,main_tally,count_bins=count_bin_edges,min_opacity=0.15,max_opacity=0.95)
    table_text = write_country_tables(main_tally,table_coulmn_header,left_col_title=table_country_coulmn_header)
    if alternate_tables_and_maps:
        html_output_text += table_text + '\n\n<br>\n\n'
        html_output_text += svg_html_code.format(map_output_filename+'.svg') + '\n\n<br>\n\n'
    else:
        html_maps_output_text += svg_html_code.format(map_output_filename+'.svg') + '\n\n<br>\n\n'
        html_tables_output_text += table_text + '\n\n<br>\n\n'

    html_output_text += html_maps_output_text + html_tables_output_text
    
    
    f2 = open(html_output_filepath,"w+", encoding='utf_8')
    f2.write(html_output_text)
    f2.close() 

    if data_is_in_spreadsheet: wb.close()
    
    
    if save_data_to_pickle and the_spreadsheet_was_parsed:
        print('saving updated pickle file...\n\t{} elapsed'.format(datetime.datetime.now() - start_time))
        with open(pickle_filepath, 'wb') as handle:
            pickle.dump(user_data_pickle, handle, protocol=pickle.HIGHEST_PROTOCOL)
    if save_data_to_pickle and USE_GUI:
        print('saving pickle file with updated settings...\n\t{} elapsed'.format(datetime.datetime.now() - start_time))
        if 'GUI_settings' in user_data_pickle[main_data_keyname]: # overwrite old data
            user_data_pickle[main_data_keyname]['GUI_settings'] = values
        else: # new entry
            user_data_pickle[main_data_keyname].update({'GUI_settings':values})
        user_data_pickle[main_data_keyname]
        with open(pickle_filepath, 'wb') as handle:
            pickle.dump(user_data_pickle, handle, protocol=pickle.HIGHEST_PROTOCOL)
    
    print('Done! '+'\n\t{} elapsed'.format(datetime.datetime.now() - start_time))
    



# ========================================================================
#                           DEFINITION OF GUI
# ========================================================================


if USE_GUI:
    
    # Get default path to file
    try:
        f=open('default_path_to_data_file.txt')
        lines=f.readlines()
        f.close()
        default_data_filepath = lines[0].strip()
    except:
        default_data_filepath = os.getcwd() + '/' +  'default.xlsx'
    if os.path.exists(default_data_filepath) and '.xls' in default_data_filepath:
        file = default_data_filepath
        wb = openpyxl.load_workbook(filename = file, read_only=True,data_only=True)  # read only is important for making it load quickly
        list_of_sheetnames = wb.sheetnames
        is_default_file_spreadsheet = True
    else:
        list_of_sheetnames = ['Sheet 1']
        is_default_file_spreadsheet = False
    
    
    default_data_folder = os.path.split(default_data_filepath)[0] #+ '/'
    default_output_folder = default_data_folder
    
    layout = [
          [sg.Text('Select the input file (Excel, .csv, or .tsv) and specify the location of the data:')],
          [sg.Text('Input file',size=(9, 1)), sg.InputText(size=(60, 1), key='data_filepath',default_text=default_data_filepath, enable_events=True), sg.FileBrowse(initial_folder=default_data_folder,tooltip='Select the file containing the country data.')],
          [sg.Text('Select sheet name (spreadsheets only)',size=(30, 1)), sg.InputCombo(list_of_sheetnames, key='select_worksheet',default_value=list_of_sheetnames[0],disabled=(not is_default_file_spreadsheet),enable_events=True)],
          [sg.Text('Is the data already tallied?'),sg.Radio('Yes', "RADIO3", key="already_tallied_yes", default=True, enable_events=True),
           sg.Radio('No', "RADIO3", key="already_tallied_no", default=False, enable_events=True)],
          [sg.Text('Note: "Yes" means each country already has an associated value assigned to it; "No" means the \n          final value associated with each country is equal to its number of appearances in the list.\n')],
          [sg.Text('# of header rows'), sg.InputText(size=(5, 1), key='n_skip_rows',default_text=n_skip_rows),
           sg.Text('Country column #'), sg.InputText(size=(5, 1), key='i_country_col',default_text=i_country_col),
           sg.Text('Tallied data column #'), sg.InputText(size=(5, 1), key='i_data_col',default_text=i_data_col,disabled=(not is_data_already_in_tallied_form))],
          [sg.Text('Note: Header rows will be skipped; column numbers start at 1 and can be Excel-style letters.\n         "Country column #" denotes where the country names are listed.\n         "Tallied data column #" denotes where already tallied data values are listed.')],
          [sg.Text('\nSelect output folder and settings for the output map, table and HTML file:')],
          [sg.Text('Output folder'), sg.InputText(size=(60, 1), key='output_folder',default_text=default_output_folder), sg.FolderBrowse(initial_folder=default_output_folder,tooltip='Select the folder where output files will be written.')],
          [sg.In("", visible=False, enable_events=True, key='set_line_color'),
           sg.Text('Select color theme for the maps'), sg.ColorChooserButton(key='custom_style_color',button_text='Pick color', target='set_line_color',button_color=('#ffffff', custom_style_color))],
          [sg.Text('Include SVG in HTML as'),sg.Radio('<embed>', "RADIO2", key="html_svg_embed", default=True),sg.Radio('<object>', "RADIO2", key="html_svg_object"),sg.Radio('<img>', "RADIO2", key="html_svg_img"),sg.Radio('<iframe>', "RADIO2", key="html_svg_iframe")],
          #[sg.Text('\nCustomize the text in the output map and table and their filenames.')],
          #[sg.Text('\nMap and table options:')],
          [sg.Text('HTML filename',size=(12, 1)), sg.InputText(size=(40, 1), key='user_html_output_filename',default_text=user_html_output_filename),sg.Text('.html')],
          [sg.Text('Map filename',size=(12, 1)), sg.InputText(size=(40, 1), key='user_map_output_filename',default_text=user_map_output_filename),sg.Text('.svg', size=(8,1)),sg.Text('',size=(1, 1))],
          [sg.Text('Map title',size=(12, 1)), sg.InputText(size=(50, 1), key='user_map_title',default_text=user_map_title),sg.Text('',size=(2, 1))],
          [sg.Text('Hover text',size=(12, 1)), sg.InputText(size=(50, 1), key='user_map_hover_text',default_text=user_map_hover_text),sg.Text('',size=(2, 1))],
          [sg.Checkbox('Show legend', default=True,key='show_legend',enable_events=True),sg.Text('Legend label'), sg.InputText(size=(20, 1), key='legend_label',default_text=legend_label),
           sg.Checkbox('Use Pygal default binning', default=False,key='use_pygal_color_bins',disabled=show_legend,enable_events=True)],
          [sg.Text('Data binning',size=(10, 1)),sg.Radio('Manual entry', "RADIO4", key="manually_enter_bins", default=True, enable_events=True),
           sg.Radio('Automatic (linear)', "RADIO4", key="automatically_generate_bins_lin", default=False, enable_events=True),
           sg.Radio('Automatic (log)', "RADIO4", key="automatically_generate_bins_log", default=False, enable_events=True)],
          [sg.Text('',size=(10, 1)),sg.Checkbox('add overflow bin', default=include_overflow_bin,key='include_overflow_bin'),
           sg.Text('# of bins ='),sg.Combo([str(i+1) for i in range(7)], key='num_auto_bins',default_value=num_auto_bins,disabled=(manually_enter_bins)),
           sg.Checkbox('force integer bin edges', default=force_integer_bin_edges,key='force_integer_bin_edges')], # disabled=(manually_enter_bins)
          [sg.Text('Map bins',size=(12, 1)), sg.InputText(size=(50, 1), key='tally_bin_edges',default_text=str(tally_bin_edges).strip('[]'),disabled=(not manually_enter_bins))],
          [sg.Text('Table left title',size=(12, 1)), sg.InputText(size=(50, 1), key='user_table_country_coulmn_header',default_text=user_table_country_coulmn_header),sg.Text('',size=(2, 1))],
          [sg.Text('Table right title',size=(12, 1)), sg.InputText(size=(50, 1), key='user_table_coulmn_header',default_text=user_table_coulmn_header),sg.Text('',size=(2, 1))],
          [sg.Button('Run',size=(15, 1)), sg.Button('Run (keep window open)',size=(25, 1)), sg.Button('Exit',size=(15, 1))]]
    
    

    window = sg.Window('Binned World Map Generator', layout,finalize=True)
    chooser = window['custom_style_color']
       
    # Event Loop to process "events" and get the "values" of the inputs
    startup_event = True
    while True:
        if startup_event:
            event = 'startup_event'
        else:
            event, values = window.read()
        
        if event == sg.WIN_CLOSED or event == 'Exit': # if user closes window or clicks cancel
            sys.exit()
        
        elif event == 'data_filepath' or event == 'select_worksheet' or event == 'startup_event':
            startup_event = False
            if event == 'startup_event':
                data_filepath = default_data_filepath
            else:
                data_filepath = values['data_filepath']
            if event == 'data_filepath':
                data_filepath = values['data_filepath']
                if os.path.isfile(data_filepath):
                    if '.xls' in data_filepath:
                        file = data_filepath
                        wb = openpyxl.load_workbook(filename = file, read_only=True,data_only=True)  # read only is important for making it load quickly
                        list_of_sheetnames = list(wb.sheetnames)
                        window['select_worksheet'].update(values=list_of_sheetnames,disabled=False)
                    else:
                        window['select_worksheet'].update(values=list_of_sheetnames,disabled=True)
                else:
                    print('Selected file path does not exist.')
                    
            if os.path.isfile(data_filepath):
                default_pickle_filepath = os.path.splitext(data_filepath)[0] + '.pickle'    
                if os.path.isfile(default_pickle_filepath):
                    if read_data_from_pickle:
                        with open(default_pickle_filepath, 'rb') as handle:
                            user_data_pickle = pickle.load(handle)
                            if '.xls' in data_filepath:
                                if event == 'select_worksheet':
                                    map_category = values['select_worksheet']
                                    main_data_keyname = map_category + '_data_all'
                                else: # default to first sheet in dict
                                    main_data_keyname = list(user_data_pickle.keys())[0] 
                            else:
                                map_category = ''
                                main_data_keyname = map_category + '_data_all'
                            if main_data_keyname in user_data_pickle.keys():
                                gui_settings = user_data_pickle[main_data_keyname]['GUI_settings']
                                for key in gui_settings:
                                    if 'Browse' in key: continue
                                    if 'set_line_color' in key: continue
                                    try:
                                        window[key].update(gui_settings[key])
                                    except:
                                        pass
                
                                
                                window['custom_style_color'].Update(button_color=('#ffffff', gui_settings['set_line_color']))
                                window['set_line_color'].Update(gui_settings['set_line_color'])
                                #if key == 'already_tallied_yes' or key == 'already_tallied_no':
                                is_data_already_in_tallied_form   = gui_settings['already_tallied_yes']
                                if not is_data_already_in_tallied_form:
                                    window['i_data_col'].update(background_color='light gray',text_color='gray')
                                else:
                                    window['i_data_col'].update(background_color='white',text_color='black')
                                window['i_data_col'].update(disabled=(not is_data_already_in_tallied_form))
                            
                                #elif key == "manually_enter_bins" or key == "automatically_generate_bins_lin" or key == "automatically_generate_bins_log":
                                manually_enter_bins = gui_settings["manually_enter_bins"]
                                automatically_generate_bins_lin = gui_settings["automatically_generate_bins_lin"]
                                automatically_generate_bins_log = gui_settings["automatically_generate_bins_log"]
                                if not manually_enter_bins:
                                    window['tally_bin_edges'].update(background_color='light gray',text_color='gray')
                                else:
                                    window['tally_bin_edges'].update(background_color='white',text_color='black')
                                window['tally_bin_edges'].update(disabled=(not manually_enter_bins))
                                window['num_auto_bins'].update(disabled=manually_enter_bins)
                                #window["force_integer_bin_edges"].update(disabled=manually_enter_bins)
                            
                                #elif key == 'show_legend':
                                show_legend = gui_settings['show_legend']
                                window['use_pygal_color_bins'].update(disabled=show_legend)
                                if not show_legend:
                                    window['legend_label'].update(background_color='light gray',text_color='gray',disabled=True)
                                else:
                                    window['legend_label'].update(background_color='white',text_color='black',disabled=False)
                            
                                if gui_settings['use_pygal_color_bins']:
                                    use_pygal_color_bins = gui_settings['use_pygal_color_bins']
                                    if (not use_pygal_color_bins and not gui_settings["manually_enter_bins"]) or use_pygal_color_bins:
                                        window['tally_bin_edges'].update(background_color='light gray',text_color='gray',disabled=True)
                                    else:
                                        window['tally_bin_edges'].update(background_color='white',text_color='black',disabled=False)
                                    if (not use_pygal_color_bins and gui_settings["manually_enter_bins"]) or use_pygal_color_bins:
                                        window['num_auto_bins'].update(disabled=True)
                                    else:
                                        window['num_auto_bins'].update(disabled=False)
                                    window["manually_enter_bins"].update(disabled=use_pygal_color_bins)
                                    window["automatically_generate_bins_lin"].update(disabled=use_pygal_color_bins)
                                    window["automatically_generate_bins_log"].update(disabled=use_pygal_color_bins)
                                    window["include_overflow_bin"].update(disabled=use_pygal_color_bins)
                                    window["force_integer_bin_edges"].update(disabled=use_pygal_color_bins)
            
                
            
    
    
        
        
        
        
        
        
        
        elif event == 'set_line_color':
            window['custom_style_color'].Update(button_color=('#ffffff', values[event]))
        
        elif event == 'already_tallied_yes' or event == 'already_tallied_no':
            is_data_already_in_tallied_form   = values['already_tallied_yes']
            if not is_data_already_in_tallied_form:
                window['i_data_col'].update(background_color='light gray',text_color='gray')
            else:
                window['i_data_col'].update(background_color='white',text_color='black')
            window['i_data_col'].update(disabled=(not is_data_already_in_tallied_form))
        
        elif event == "manually_enter_bins" or event == "automatically_generate_bins_lin" or event == "automatically_generate_bins_log":
            manually_enter_bins = values["manually_enter_bins"]
            automatically_generate_bins_lin = values["automatically_generate_bins_lin"]
            automatically_generate_bins_log = values["automatically_generate_bins_log"]
            if not manually_enter_bins:
                window['tally_bin_edges'].update(background_color='light gray',text_color='gray')
            else:
                window['tally_bin_edges'].update(background_color='white',text_color='black')
            window['tally_bin_edges'].update(disabled=(not manually_enter_bins))
            window['num_auto_bins'].update(disabled=manually_enter_bins)
            window["force_integer_bin_edges"].update(disabled=False)
        
        elif event == 'show_legend':
            show_legend = values['show_legend']
            window['use_pygal_color_bins'].update(disabled=show_legend)
            if not show_legend:
                window['legend_label'].update(background_color='light gray',text_color='gray',disabled=True)
            else:
                window['legend_label'].update(background_color='white',text_color='black',disabled=False)
        
        elif event == 'use_pygal_color_bins':
            use_pygal_color_bins = values['use_pygal_color_bins']
            if (not use_pygal_color_bins and not values["manually_enter_bins"]) or use_pygal_color_bins:
                window['tally_bin_edges'].update(background_color='light gray',text_color='gray',disabled=True)
            else:
                window['tally_bin_edges'].update(background_color='white',text_color='black',disabled=False)
            if (not use_pygal_color_bins and values["manually_enter_bins"]) or use_pygal_color_bins:
                window['num_auto_bins'].update(disabled=True)
            else:
                window['num_auto_bins'].update(disabled=False)
            window["manually_enter_bins"].update(disabled=use_pygal_color_bins)
            window["automatically_generate_bins_lin"].update(disabled=use_pygal_color_bins)
            window["automatically_generate_bins_log"].update(disabled=use_pygal_color_bins)
            window["include_overflow_bin"].update(disabled=use_pygal_color_bins)
            window["force_integer_bin_edges"].update(disabled=use_pygal_color_bins)
        
        
        
        
        
        
        
        
        
        
        
        
        elif event == 'Run' or event=='Run (keep window open)':
            #print('You entered ', values)
            # values['']
            
            data_filepath = values['data_filepath']
            output_folder = values['output_folder'] + '/'
            data_filename = os.path.basename(data_filepath)
            if '\\' in output_folder: output_folder = output_folder.replace('\\','/')
            if '.xls' in data_filename:
                data_is_in_spreadsheet = True
            else:
                data_is_in_spreadsheet = False
                if os.path.splitext(data_filepath)[1] not in ['.csv','.tsv','.txt']:
                    print('Selected data file format unknown, will assume it is a .csv file.')
                    filetype = 'csv'
                elif os.path.splitext(data_filepath)[1] == '.tsv':
                    filetype = 'tsv'
                else:
                    filetype = 'csv'
            
            wbname                          = values['select_worksheet']
            n_skip_rows                     = int(values['n_skip_rows'])
            is_data_already_in_tallied_form = values['already_tallied_yes']
            i_country_col                   = values['i_country_col']
            i_data_col                      = values['i_data_col']
            
            if values['set_line_color']!='': # change from default value
                custom_style_color = values['set_line_color']
            
            legend_label = values['legend_label']
            show_legend  = values['show_legend']
            
            user_html_output_filename   = values['user_html_output_filename']
            user_map_output_filename    = values['user_map_output_filename']
            user_map_title              = values['user_map_title']
            user_map_hover_text         = values['user_map_hover_text']
            user_tallied_quantity       = user_map_hover_text
            user_table_coulmn_header    = values['user_table_coulmn_header']
            user_table_country_coulmn_header = values['user_table_country_coulmn_header']
            
            if values['html_svg_embed']:
                html_include_map_style = html_include_map_style_options[0]
            elif values['html_svg_object']:
                html_include_map_style = html_include_map_style_options[1]
            elif values['html_svg_img']:
                html_include_map_style = html_include_map_style_options[2]
            elif values['html_svg_iframe']:
                html_include_map_style = html_include_map_style_options[3]
            else:
                html_include_map_style = html_include_map_style_options[0]
            
            use_pygal_color_bins = values["use_pygal_color_bins"]
            manually_enter_bins = values["manually_enter_bins"]
            automatically_generate_bins_lin = values["automatically_generate_bins_lin"]
            automatically_generate_bins_log = values["automatically_generate_bins_log"]
            num_auto_bins = int(values["num_auto_bins"])
            force_integer_bin_edges = values["force_integer_bin_edges"]
            
            tally_bin_edges                 = eval('['+values['tally_bin_edges']+']')
            include_overflow_bin            = values["include_overflow_bin"]
            
            if not os.path.isfile(data_filepath):
                print('Input file does not exist at specified location:\n\t'+data_filepath+'\nQuitting program.')
                sys.exit()
                
            pickle_filepath = os.path.splitext(data_filepath)[0] + '.pickle'
            
            if event=='Run (keep window open)':
                sg.Print('Program is running, please wait...', do_not_reroute_stdout=False)
                generate_maps_and_tables()
            else:
                break
    
    window.close()

generate_maps_and_tables()


